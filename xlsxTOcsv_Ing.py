# -*- coding: utf-8 -*-
"""
Created on Fri Dec  4 11:17:50 2020

@author: ppa16, dsantos
"""
import os
from xlsx2csv import Xlsx2csv
from openpyxl import load_workbook
import re

proyecto = 'C:\\Users\\eppa\\Documents\\Home_Office\\PROYECCION\\02. Ronda'

os.getcwd()
os.chdir(proyecto)
os.listdir()

#%% Changing extension xlsx to csv

files = os.listdir('.\\Inputs\\xlsx')
for archivo in files:
    NSheet = 0
    a = re.search('.xlsx$', archivo)
    if a is not None:
        LasSheets = load_workbook('.\\Inputs\\xlsx\\'+archivo, read_only=True).sheetnames
        for s in LasSheets:
            NSheet = NSheet + 1
            print('Tranformando hoja ', s, ' del archivo ', archivo)
            Xlsx2csv('.\\Inputs\\xlsx\\'+archivo, outputencoding="utf-8").convert('.\\Inputs\\csv\\'+(archivo.replace('.xlsx', '_s'+str(NSheet) + '.csv')), sheetid = NSheet)


